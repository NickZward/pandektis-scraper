#!/usr/bin/env python3
"""
Pandektis Scraper v4 — Correct HTML Selectors
===============================================
Previous versions failed because they looked for DSpace's standard
"itemDisplayTable" class, but Pandektis uses a custom theme with its
own CSS classes. This version uses the correct selectors, confirmed
by live browser inspection:

  div.ekt_met_item        → one metadata row (label + value pairs)
  div.ekt_met_curved      → label  e.g. "Παλαιά ονομασία :"
  div.ekt_met_curved_metadata → value e.g. "21η Απριλίου"
  div.ekt_met_metadata_title  → page title (old -- new)

The site is plain server-side HTML — no JavaScript rendering needed.
The scraper uses requests + BeautifulSoup directly.

Field mapping (confirmed from live page):
  "Παλαιά ονομασία"        → old_name_gr
  "Old name"               → old_name_en  (site provides English!)
  "Νέα ονομασία"           → new_name_gr
  "New name"               → new_name_en  (site provides English!)
  "Νομός"                  → prefecture_gr
  "Prefecture"             → prefecture_en
  "Επαρχία"                → province_gr
  "Province"               → province_en
  "Ημερομηνία μετονομασίας"→ renaming_date

Usage:
    python pandektis_scraper_v4.py

    To run the full 4413-record job:
        python pandektis_scraper_v4.py --full

Requirements:
    pip install requests beautifulsoup4 lxml openpyxl geopy tenacity tqdm
"""

import argparse
import json
import time
import logging
import re
import sys
from datetime import datetime
from typing import Optional, Dict, List, Any, Tuple

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(it, **kw): return it

# ─── Config ───────────────────────────────────────────────────────────────────

BASE_URL       = "https://pandektis.ekt.gr/pandektis"
COLLECTION_URL = (
    f"{BASE_URL}/handle/10442/4968/browse"
    "?type=title&sort_by=1&order=ASC&rpp={rpp}"
)

REQUEST_DELAY  = 0.8    # seconds between item page fetches
GEOCODE_DELAY  = 1.2    # seconds between Nominatim calls (ToS requirement)
REQUEST_TIMEOUT = 25

# ─── Logging ──────────────────────────────────────────────────────────────────

def setup_logging(log_file: str):
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )

log = logging.getLogger(__name__)

# ─── HTTP session ─────────────────────────────────────────────────────────────

def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "el-GR,el;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
    })
    retry = Retry(total=4, backoff_factor=2,
                  status_forcelist=[429, 500, 502, 503, 504])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    return s

session = make_session()


def fetch_soup(url: str) -> BeautifulSoup:
    resp = session.get(url, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    resp.encoding = "utf-8"
    return BeautifulSoup(resp.text, "lxml")

# ─── Step 1: collect item URLs from browse page ───────────────────────────────

def get_item_urls(limit: int) -> List[str]:
    url = COLLECTION_URL.format(rpp=limit)
    log.info(f"Fetching browse page (rpp={limit})…")
    soup = fetch_soup(url)

    urls = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if re.search(r"/handle/\d+/\d+$", href) and "4968" not in href:
            # hrefs are root-relative: /pandektis/handle/...
            full = "https://pandektis.ekt.gr" + href if href.startswith("/") else href
            if full not in urls:
                urls.append(full)

    log.info(f"Found {len(urls)} item URLs.")
    return urls[:limit]

# ─── Step 2: parse a single item page ────────────────────────────────────────

def clean(s: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def parse_date(raw: str) -> str:
    raw = clean(raw)
    for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%Y"]:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return raw


def extract_metadata(soup: BeautifulSoup, url: str) -> Dict[str, Any]:
    """
    Parse a Pandektis item page using the confirmed CSS selectors:

      div.ekt_met_item        — one row of metadata
        div.ekt_met_curved          — label(s) in that row
        div.ekt_met_curved_metadata — value(s) in that row

    Each ekt_met_item contains two label+value pairs side by side:
      Greek label | Greek value | English label | English value
    """
    record: Dict[str, str] = {
        "old_name_gr":    "",
        "old_name_en":    "",
        "new_name_gr":    "",
        "new_name_en":    "",
        "prefecture_gr":  "",
        "prefecture_en":  "",
        "province_gr":    "",
        "province_en":    "",
        "renaming_date":  "",
        "source_url":     url,
    }

    for item_div in soup.find_all("div", class_="ekt_met_item"):
        labels = [clean(d.get_text()) for d in item_div.find_all("div", class_="ekt_met_curved")]
        values = [clean(d.get_text()) for d in item_div.find_all("div", class_="ekt_met_curved_metadata")]

        # Build label→value pairs
        pairs = list(zip(labels, values))
        for label, value in pairs:
            label_norm = label.rstrip(":").strip().lower()

            if "παλαιά ονομασία" in label_norm:
                record["old_name_gr"] = value
            elif "old name" in label_norm:
                record["old_name_en"] = value
            elif "νέα ονομασία" in label_norm:
                record["new_name_gr"] = value
            elif "new name" in label_norm:
                record["new_name_en"] = value
            elif any(k in label_norm for k in ["νομ"]) and "ονομ" not in label_norm and "prefecture" not in label_norm:
                record["prefecture_gr"] = value
            elif "prefecture" in label_norm:
                record["prefecture_en"] = value
            elif "παρχ" in label_norm and "province" not in label_norm:
                record["province_gr"] = value
            elif "province" in label_norm:
                record["province_en"] = value
            elif "ημερομηνία μετονομασίας" in label_norm or "date of renaming" in label_norm:
                if not record["renaming_date"]:
                    record["renaming_date"] = parse_date(value)

    return record

# ─── Step 3: geocode ──────────────────────────────────────────────────────────

geolocator = Nominatim(user_agent="pandektis-scraper-v4/1.0", timeout=12)
_geocache: Dict[str, Optional[Tuple[float, float]]] = {}


def geocode(new_name_gr: str, new_name_en: str, prefecture_en: str) -> Optional[Tuple[float, float]]:
    """
    Try geocoding with multiple query strategies, most to least specific.
    Uses both English and Greek names, with and without prefecture.
    Also tries stripping common suffixes for better matching.
    """
    def variants(name: str) -> list:
        """Generate name variants by stripping common Greek suffixes."""
        v = [name]
        for suffix in ["ον", "ιον", "αιον", "ειον", "ιο", "αι"]:
            if name.lower().endswith(suffix) and len(name) > len(suffix) + 3:
                v.append(name[:-len(suffix)])
        return v

    candidates = []
    # Most specific: English name + prefecture
    if new_name_en and prefecture_en:
        candidates.append(f"{new_name_en}, {prefecture_en}, Greece")
    # English name alone
    if new_name_en:
        candidates.append(f"{new_name_en}, Greece")
    # Greek name + prefecture
    if new_name_gr and prefecture_en:
        candidates.append(f"{new_name_gr}, {prefecture_en}, Greece")
        # Variants of Greek name + prefecture
        for v in variants(new_name_gr)[1:]:
            candidates.append(f"{v}, {prefecture_en}, Greece")
    # Greek name alone + variants
    if new_name_gr:
        candidates.append(f"{new_name_gr}, Greece")

    for query in candidates:
        key = query.lower().strip()
        if key in _geocache:
            result = _geocache[key]
            if result:
                return result
            continue
        try:
            time.sleep(GEOCODE_DELAY)
            loc = geolocator.geocode(query, exactly_one=True,
                                     language="en", country_codes="gr")
            result = (round(loc.longitude, 6), round(loc.latitude, 6)) if loc else None
            _geocache[key] = result
            if result:
                log.debug(f"  Geocoded '{query}' → {result}")
                return result
        except (GeocoderTimedOut, GeocoderServiceError) as e:
            log.warning(f"  Geocoder error for '{query}': {e}")
            _geocache[key] = None

    return None

# ─── Step 4: save outputs ─────────────────────────────────────────────────────

def save_geojson(records: List[Dict], path: str):
    features = []
    for r in records:
        coords = r.get("_coords")
        props = {k: v for k, v in r.items() if not k.startswith("_")}
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": list(coords),
            } if coords else None,
            "properties": props,
        })
    fc = {
        "type": "FeatureCollection",
        "name": "Greek Settlement Renamings — Pandektis/EKT",
        "features": features,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(fc, f, ensure_ascii=False, indent=2)
    log.info(f"✓ GeoJSON saved → {path}")


def save_xlsx(records: List[Dict], path: str, label: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settlements"

    COLS = [
        ("old_name_gr",   "Old Name (Greek)"),
        ("old_name_en",   "Old Name (English)"),
        ("new_name_gr",   "New Name (Greek)"),
        ("new_name_en",   "New Name (English)"),
        ("prefecture_gr", "Prefecture (Greek)"),
        ("prefecture_en", "Prefecture (English)"),
        ("province_gr",   "Province (Greek)"),
        ("province_en",   "Province (English)"),
        ("renaming_date", "Renaming Date"),
        ("longitude",     "Longitude"),
        ("latitude",      "Latitude"),
        ("geocoded",      "Geocoded?"),
        ("source_url",    "Source URL"),
    ]

    hf = PatternFill("solid", fgColor="1F4E79")
    hn = Font(bold=True, color="FFFFFF", size=11)
    af = PatternFill("solid", fgColor="D6E4F0")

    for ci, (_, lbl) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=lbl)
        c.fill = hf; c.font = hn
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for ri, rec in enumerate(records, 2):
        coords = rec.get("_coords")
        vals = {
            "old_name_gr":   rec.get("old_name_gr", ""),
            "old_name_en":   rec.get("old_name_en", ""),
            "new_name_gr":   rec.get("new_name_gr", ""),
            "new_name_en":   rec.get("new_name_en", ""),
            "prefecture_gr": rec.get("prefecture_gr", ""),
            "prefecture_en": rec.get("prefecture_en", ""),
            "province_gr":   rec.get("province_gr", ""),
            "province_en":   rec.get("province_en", ""),
            "renaming_date": rec.get("renaming_date", ""),
            "longitude":     coords[0] if coords else "",
            "latitude":      coords[1] if coords else "",
            "geocoded":      "Yes" if coords else "No",
            "source_url":    rec.get("source_url", ""),
        }
        fill = af if ri % 2 == 0 else None
        for ci, (key, _) in enumerate(COLS, 1):
            cell = ws.cell(row=ri, column=ci, value=vals[key])
            if fill: cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    widths = [26, 26, 26, 26, 22, 22, 20, 20, 14, 14, 14, 11, 52]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total     = len(records)
    geocoded  = sum(1 for r in records if r.get("_coords"))
    non_empty = sum(1 for r in records if r.get("old_name_gr"))
    ws2["A1"] = f"Pandektis — {label}"
    ws2["A1"].font = Font(bold=True, size=13)
    for row, (lbl, val) in enumerate([
        ("Total records",        total),
        ("Records with data",    non_empty),
        ("Geocoded",             geocoded),
        ("Geocode success rate", f"{geocoded/total*100:.1f}%" if total else "N/A"),
        ("Export date",          datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source",               COLLECTION_URL.format(rpp=total)),
    ], start=3):
        ws2.cell(row=row, column=1, value=lbl).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=val)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 60

    wb.save(path)
    log.info(f"✓ XLSX saved  → {path}")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Pandektis settlement renaming scraper")
    parser.add_argument("--full", action="store_true",
                        help="Scrape all 4413 records (takes ~3 hours)")
    args = parser.parse_args()

    limit       = 4413 if args.full else 100
    label       = f"Full Extract ({limit} records)" if args.full else f"Test Extract ({limit} records)"
    log_file    = "pandektis_full.log" if args.full else "pandektis_test_100.log"
    geojson_out = "pandektis_settlements.geojson" if args.full else "pandektis_test_100.geojson"
    xlsx_out    = "pandektis_settlements.xlsx"    if args.full else "pandektis_test_100.xlsx"

    setup_logging(log_file)

    log.info("━" * 55)
    log.info(f"  Pandektis Scraper v4 — {label}")
    log.info("━" * 55)

    # ── 1. Collect URLs ───────────────────────────────────────────────────────
    urls = get_item_urls(limit)
    if not urls:
        log.error("No URLs found. Check your internet connection.")
        return

    # ── 2. Extract metadata ───────────────────────────────────────────────────
    records = []
    failed  = []

    log.info(f"\nExtracting metadata from {len(urls)} pages…")
    for i, url in enumerate(tqdm(urls, desc="Extracting", unit="record"), 1):
        try:
            soup = fetch_soup(url)
            rec  = extract_metadata(soup, url)
            records.append(rec)
            log.info(
                f"  [{i:4d}] "
                f"{rec['old_name_gr'] or '?':30s} → "
                f"{rec['new_name_gr'] or '?':30s}  "
                f"[{rec['prefecture_en'] or rec['prefecture_gr'] or 'no prefecture'}]"
            )
        except Exception as e:
            log.warning(f"  FAILED {url}: {e}")
            failed.append(url)
        time.sleep(REQUEST_DELAY)

    non_empty = sum(1 for r in records if r.get("old_name_gr"))
    log.info(f"\nExtracted {len(records)} records ({non_empty} with data, {len(failed)} failed).")

    if non_empty == 0:
        log.error(
            "All records are empty — the page structure may have changed.\n"
            "Please send the log file for diagnosis."
        )
        return

    # ── 3. Geocode ────────────────────────────────────────────────────────────
    log.info("\nGeocoding (using new English name + English prefecture)…")
    for rec in tqdm(records, desc="Geocoding", unit="record"):
        rec["_coords"] = geocode(
            rec.get("new_name_gr", ""),
            rec.get("new_name_en", ""),
            rec.get("prefecture_en", ""),
        )

    geocoded = sum(1 for r in records if r.get("_coords"))
    log.info(f"Geocoded {geocoded}/{len(records)} ({geocoded/len(records)*100:.0f}% success).")

    # ── 4. Save ───────────────────────────────────────────────────────────────
    save_geojson(records, geojson_out)
    save_xlsx(records, xlsx_out, label)

    log.info("\n" + "─" * 55)
    log.info("DONE.")
    log.info(f"  GeoJSON : {geojson_out}")
    log.info(f"  XLSX    : {xlsx_out}")
    log.info(f"  Log     : {log_file}")
    if failed:
        log.info(f"\n  Failed URLs ({len(failed)}):")
        for u in failed:
            log.info(f"    {u}")

    if not args.full:
        log.info(
            "\n  ✓ Test complete! If data looks good, run the full job:\n"
            "      python pandektis_scraper_v4.py --full\n"
            "  (estimated time: ~2-3 hours for 4413 records)"
        )


if __name__ == "__main__":
    main()
